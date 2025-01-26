from openpyxl import load_workbook


def input_data(path,sheet_name):
    load_wk= load_workbook(path)
    sheet=load_wk[sheet_name]
    data = []
    for rows in sheet.iter_rows(min_row=2,values_only=True):
        data.append(rows)
    return data
