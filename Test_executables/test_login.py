from COnfest.data_xl import input_data
import time
import openpyxl
from selenium import webdriver
import pytest

from Test_cases.login_ import TestLogin


data_i = input_data("/home/aryan/Documents/btes/XLSS.xlsx","Sheet1")
@pytest.fixture(scope="class")
def driver():
    driver=webdriver.Firefox()
    driver.get("http://www.automationpractice.pl/index.php?controller=authentication&back=my-account#account-creation")
    yield driver
    driver.quit()


@pytest.mark.parametrize("email, password,data_write",data_i)
def test_login_functionality(email, password, data_write, driver, i=1):
    l = TestLogin(driver)
    l.login_in(email,password)
    l.login_button()
    expected = driver.title
    assert driver.title == expected,openpyxl.load_workbook("/home/aryan/Documents/btes/XLSS.xlsx")["Sheet1"].cell(row=i, column=3,value="Fail")
    openpyxl.load_workbook("/home/aryan/Documents/btes/XLSS.xlsx")["Sheet1"].cell(row=i, column=3,value="Pass")
    time.sleep(4)
    i+=1
    openpyxl.load_workbook("/home/aryan/Documents/btes/XLSS.xlsx").save("/home/aryan/Documents/btes/XLSS.xlsx")
